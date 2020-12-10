# Webdriver_auto
from assert_data.assert_result import Assert
from error_screenshot.error_web_screenshot import Screenshoot_web
from flow_auto_path.local_element_path import Flow_element
from read_writer_excel.option_excel import *
import time
from log.Log import info
from openpyxl import load_workbook
class test_path:
    def __init__(self,driver):
        self.driver = driver

        self.FLow_element = Flow_element(self.driver)

        self.Readexcel_element_path=Readexcel_element_path()
        self.Readexcel_data =Readexcel_data()
        self.document_path = Document_path()

        self.element_path = self.document_path.read_element_path()#读取文档row值
        self.case_path = self.document_path.read_data_path()
        self.element_wb = load_workbook(self.element_path)
        self.case_wb = load_workbook(self.case_path)
        self.element_max = self.element_wb['Sheet1'].max_row
        self.case_max = self.case_wb['Sheet1'].max_row
        #print(self.element_max,self.case_max)

        info("执行中")
        self.login()

    def login(self):
        for element_row in range(2,self.element_max+1):
            test_path_item = self.Readexcel_element_path.test_items(element_row)#执行功能项
            execute = self.Readexcel_element_path.elemen_execute(element_row)#是否执行
            if execute == 'N':
                continue
            elif execute == 'Y':
                for case_row in range(2,self.case_max+1):

                    test_case_item = self.Readexcel_data.test_items(case_row)#case功能项
                    if test_path_item ==test_case_item:


                        path = self.Readexcel_element_path.element_path(element_row)#拿取定位方法元素操作
                        # path_test_items = self.Readexcel_element_path.test_items(element_row)#测试功能项
                        data = self.Readexcel_data.test_data(case_row)#测试数据


                        assert_local_method = self.Readexcel_data.test_assert(case_row)#断言定位
                        assert_method = self.Readexcel_data.assert_method(case_row)#断言方法
                        assert_data = self.Readexcel_data.test_assert_data(case_row)#断言数据
                        except_result = self.Readexcel_data.test_except_result(case_row)#预期结果
                        a = 0

                        for i in path:


                            ele_path = i.split(':')
                            self.FLow_element.location(ele_path[0],ele_path[1])

                            if ele_path[2] == 'send_keys' and data !=None:#send_keys执行if，其它执行else
                                if a == len(data):
                                    a = 0
                                else:
                                    self.FLow_element.operation(send_operation=ele_path[2],data=data[a])
                                    a+=1
                            else:
                                self.FLow_element.operation(send_operation=ele_path[2])
                        time.sleep(1)
                        # print(assert_local_method)
                        self.FLow_element.location( assert_local_method[0], assert_local_method[1])

                        self.assert_data_ = self.FLow_element.operation(send_operation= assert_local_method[2])
                        try:
                            b=0
                            for assert_methodi in assert_method:
                                # print(self.assert_data_,assert_methodi,assert_data[b])
                                Assert_result = Assert(self.assert_data_[0],assert_data[b])
                                if assert_methodi == '相等':
                                    Assert_result.assert_equation()
                                elif assert_methodi == '不相等':
                                    Assert_result.assert_not_equation()
                                elif assert_methodi == '大于':
                                    Assert_result.assert_greater()
                                elif assert_methodi == '小于':
                                    Assert_result.assert_less()
                                elif assert_methodi == "in":
                                    Assert_result.assert_in()
                                elif assert_methodi == "notin":
                                    Assert_result.assert_notin()
                                else:
                                    info("无有效断言")
                                b += 1


                            info(except_result)
                        except Exception as e:
                            screen =Screenshoot_web(self.assert_data_[1])
                            screen.screenshoot()
                            info("%s执行失败"%test_path_item)
